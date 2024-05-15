import tkinter as tk
from tkinter import messagebox
import cv2
import os
import pickle
from PIL import Image, ImageTk
import openpyxl

class FaceRecognitionUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Face Recognition UI")

        # Initialize Excel workbook and sheets
        self.face_data_workbook = openpyxl.Workbook()
        self.face_data_sheet = self.face_data_workbook.active
        self.face_data_sheet.title = "Face Data"
        self.face_data_sheet.append(["Name", "Image Path"])

        self.attendance_workbook = openpyxl.Workbook()
        self.attendance_sheet = self.attendance_workbook.active
        self.attendance_sheet.title = "Attendance"

        # Initialize GUI components
        self.video_frame = tk.Label(master)
        self.video_frame.pack(side=tk.LEFT, padx=10, pady=10)

        self.name_label = tk.Label(master, text="")
        self.name_label.pack()

        self.add_face_button = tk.Button(master, text="Add Face", command=self.open_add_face_window)
        self.add_face_button.pack(pady=(0, 10))

        self.delete_face_button = tk.Button(master, text="Delete Face", command=self.open_delete_face_window)
        self.delete_face_button.pack(pady=10)

        self.check_face_data_button = tk.Button(master, text="Check Face Data", command=self.check_face_data)
        self.check_face_data_button.pack(pady=10)

        self.check_attendance_button = tk.Button(master, text="Check Attendance", command=self.check_attendance)
        self.check_attendance_button.pack(pady=10)

        # Initialize video capture and face detection
        self.video = cv2.VideoCapture(0)
        self.facedetect = cv2.CascadeClassifier('C:/Users/1233/Desktop/coding/AI project/data/haarcascade_frontalface_default.xml')

        # Load pre-trained LBPH face recognizer
        self.face_recognizer = cv2.face.createLBPHFaceRecognizer()
        self.face_recognizer.read("trained_model.yml")  # Load the trained model

        # Dictionary mapping label indices to names
        self.label_to_name = {0: "John Doe", 1: "Jane Smith"}  # Update with your own mappings

        self.display_video()

    def open_add_face_window(self):
        AddFaceWindow(self.master, self.face_data_sheet, self.face_data_workbook)

    def open_delete_face_window(self):
        DeleteFaceWindow(self.master, self.face_data_sheet)

    def check_face_data(self):
        pass  # You can implement this if needed

    def check_attendance(self):
        # Create a new window to display attendance
        attendance_window = tk.Toplevel(self.master)
        attendance_window.title("Attendance Records")

        # Load attendance data from the Excel file
        attendance_workbook = openpyxl.load_workbook("attendance.xlsx")
        attendance_sheet = attendance_workbook.active

        # Create a text widget to display attendance records
        attendance_text = tk.Text(attendance_window, width=50, height=20)
        attendance_text.pack(padx=10, pady=10)

        # Iterate through the rows and display attendance records
        for row in attendance_sheet.iter_rows(values_only=True):
            name = row[0]
            attendance_text.insert(tk.END, f"{name}\n")

        # Disable text editing
        attendance_text.config(state="disabled")

        # Close the workbook
        attendance_workbook.close()

    def display_video(self):
        ret, frame = self.video.read()
        if ret:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.facedetect.detectMultiScale(gray, 1.3, 5)
            
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
                # Get the name associated with the recognized face
                name = self.get_name_from_face(gray[y:y+h, x:x+w])
                if name:
                    self.name_label.config(text=f"Name: {name}")
                    self.mark_attendance(name)
                    
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            frame = cv2.resize(frame, (640, 480))
            img = Image.fromarray(frame)
            img = ImageTk.PhotoImage(image=img)
            self.video_frame.config(image=img)
            self.video_frame.img = img
            self.video_frame.after(10, self.display_video)
        else:
            messagebox.showerror("Error", "Failed to capture video.")

    def get_name_from_face(self, face_image):
        for (x, y, w, h) in faces:
            # Recognize face
            label, confidence = self.face_recognizer.predict(face_image)

            # If confidence is below a certain threshold, consider it a match
            if confidence < 100:
                name = self.label_to_name.get(label, "Unknown")
                return name

        return None

    def mark_attendance(self, name):
        # Mark attendance in the Excel file
        self.attendance_sheet.append([name])
        self.attendance_workbook.save("attendance.xlsx")

class AddFaceWindow:
    def __init__(self, master, face_data_sheet, face_data_workbook):
        self.master = master
        self.master.withdraw()
        self.add_face_window = tk.Toplevel(self.master)
        self.add_face_window.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.add_face_window.geometry("300x150")
        self.add_face_window.title("Add New Face")

        self.face_data_sheet = face_data_sheet
        self.face_data_workbook = face_data_workbook

        self.label_name = tk.Label(self.add_face_window, text="Enter Name:")
        self.label_name.grid(row=0, column=0, padx=10, pady=10)

        self.entry_name = tk.Entry(self.add_face_window)
        self.entry_name.grid(row=0, column=1, padx=10, pady=10)

        self.capture_face_button = tk.Button(self.add_face_window, text="Capture Face", command=self.capture_face)
        self.capture_face_button.grid(row=1, column=0, padx=10, pady=10)

        self.add_face_button = tk.Button(self.add_face_window, text="Add Face", command=self.add_face, state="disabled")
        self.add_face_button.grid(row=1, column=1, padx=10, pady=10)

    def on_closing(self):
        self.master.deiconify()
        self.add_face_window.destroy()

    def capture_face(self):
        # Placeholder for face capture logic
        messagebox.showinfo("Capture", "Face captured successfully. Please add a name and click 'Add Face'.")
        self.add_face_button.config(state="normal")

    def add_face(self):
        name = self.entry_name.get()
        if not name:
            messagebox.showerror("Error", "Please enter a name.")
            return

        # Save image path and name to the Excel sheet
        image_path = f"path_to_save/{name}.jpg"  # Replace with actual path
        self.face_data_sheet.append([name, image_path])
        self.face_data_workbook.save("face_data.xlsx")

        messagebox.showinfo("Success", "Face added successfully.")
        self.on_closing()

class DeleteFaceWindow:
    def __init__(self, master, face_data_sheet):
        self.master = master
        self.master.withdraw()
        self.delete_face_window = tk.Toplevel(self.master)
        self.delete_face_window.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.delete_face_window.geometry("300x100")
        self.delete_face_window.title("Delete Face")

        self.face_data_sheet = face_data_sheet

        self.label_name = tk.Label(self.delete_face_window, text="Enter Name:")
        self.label_name.grid(row=0, column=0, padx=10, pady=10)

        self.entry_name = tk.Entry(self.delete_face_window)
        self.entry_name.grid(row=0, column=1, padx=10, pady=10)

        self.delete_face_button = tk.Button(self.delete_face_window, text="Delete Face", command=self.delete_face)
        self.delete_face_button.grid(row=1, columnspan=2, padx=10, pady=10)

    def on_closing(self):
        self.master.deiconify()
        self.delete_face_window.destroy()

    def delete_face(self):
        name = self.entry_name.get()
        if not name:
            messagebox.showerror("Error", "Please enter a name.")
            return

        for row in range(2, self.face_data_sheet.max_row + 1):
            if self.face_data_sheet.cell(row, 1).value == name:
                self.face_data_sheet.delete_rows(row)
                break
        else:
            messagebox.showerror("Error", "Name not found.")
            return

        self.face_data_workbook.save("face_data.xlsx")
        messagebox.showinfo("Success", "Face deleted successfully.")
        self.on_closing()

def main():
    root = tk.Tk()
    app = FaceRecognitionUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
